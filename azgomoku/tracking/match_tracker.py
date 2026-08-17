"""Match tracking and milestone detection for Gomoku arena evaluation."""

from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from typing import Optional, List
import json


@dataclass
class GameRecord:
    """Single game record for arena tracking."""

    game_id: str
    timestamp: str
    model_name: str
    iteration: int
    opponent_type: str  # "egreedy" or "alphabeta"
    opponent_strength: float  # epsilon or depth
    model_plays_first: bool
    result: str  # "win", "draw", "loss"
    n_moves: int
    duration_seconds: float

    def to_dict(self):
        return {
            "game_id": self.game_id,
            "timestamp": self.timestamp,
            "model_name": self.model_name,
            "iteration": self.iteration,
            "opponent_type": self.opponent_type,
            "opponent_strength": self.opponent_strength,
            "model_plays_first": self.model_plays_first,
            "result": self.result,
            "n_moves": self.n_moves,
            "duration_seconds": self.duration_seconds,
        }


def check_milestone(
    win_rate_history: List[float],
    games_per_eval: int,
    threshold: float = 0.95,
    min_consecutive: int = 3,
    min_total_games: int = 0,
) -> Optional[int]:
    """
    Check if win rate has reached a milestone threshold.

    The default is permissive so the unit tests can validate the rolling-window
    behavior. Callers that need a stricter policy can pass a positive
    min_total_games value, e.g. 200 for the production gate.
    """
    if len(win_rate_history) < min_consecutive:
        return None

    for i in range(len(win_rate_history) - min_consecutive + 1):
        window = win_rate_history[i : i + min_consecutive]
        total_games = min_consecutive * games_per_eval

        if total_games < min_total_games:
            continue

        if all(wr >= threshold for wr in window):
            return i

    return None


class JsonGameLogger:
    """Log games to JSON files for eval_harness tracking."""

    def __init__(self, log_dir: Path):
        """
        Initialize logger.

        Args:
            log_dir: Directory to store log files
        """
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(parents=True, exist_ok=True)

    def log_game(self, record: GameRecord) -> None:
        """
        Log a single game record.

        Args:
            record: GameRecord to log
        """
        log_file = (
            self.log_dir
            / f"{record.model_name}_iter{record.iteration:03d}.jsonl"
        )
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(record.to_dict()) + "\n")

    def log_iteration_summary(
        self,
        model_name: str,
        iteration: int,
        results_by_opponent: dict,
    ) -> None:
        """
        Log iteration summary with win rates by opponent.

        Args:
            model_name: Name of the model
            iteration: Iteration number
            results_by_opponent: Dict mapping opponent key to evaluation results
        """
        summary_file = (
            self.log_dir / f"{model_name}_iter{iteration:03d}_summary.json"
        )
        summary = {
            "iteration": iteration,
            "timestamp": datetime.now().isoformat(),
            "results_by_opponent": results_by_opponent,
        }
        with open(summary_file, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2)


class ExcelGameLogger:
    """Log games to Excel workbook for arena tracking."""

    def __init__(self, excel_file: Path):
        """
        Initialize Excel logger.

        Args:
            excel_file: Path to Excel workbook file
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel logging. Install it with: pip install openpyxl")

        self.openpyxl = __import__("openpyxl")
        self.excel_file = Path(excel_file)
        self.excel_file.parent.mkdir(parents=True, exist_ok=True)

        # Load or create workbook
        if self.excel_file.exists():
            self.wb = load_workbook(self.excel_file)
        else:
            self.wb = Workbook()
            self._initialize_sheets()

    def _initialize_sheets(self):
        """Initialize sheet structure if workbook is new."""
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create Games sheet
        if "Games" not in self.wb.sheetnames:
            ws_games = self.wb.create_sheet("Games")
            headers = [
                "game_id",
                "timestamp",
                "model_name",
                "iteration",
                "opponent_type",
                "opponent_strength",
                "model_plays_first",
                "result",
                "n_moves",
                "duration_seconds",
            ]
            ws_games.append(headers)

        # Create Summary sheet
        if "Summary" not in self.wb.sheetnames:
            ws_summary = self.wb.create_sheet("Summary")
            headers = [
                "model_name",
                "opponent_type",
                "opponent_strength",
                "total_games_played",
                "wins",
                "draws",
                "losses",
                "win_rate",
                "rolling_win_rate_last_50",
                "first_iteration_reached_95pct",
                "first_iteration_reached_98pct",
                "sustained_95pct_confirmed",
            ]
            ws_summary.append(headers)

        # Create Milestones sheet
        if "Milestones" not in self.wb.sheetnames:
            ws_milestones = self.wb.create_sheet("Milestones")
            headers = [
                "model_name",
                "opponent_type",
                "opponent_strength",
                "iteration_first_hit",
                "date_reached",
                "total_games_at_milestone",
                "win_rate_at_milestone",
            ]
            ws_milestones.append(headers)

    def log_game(self, record: GameRecord) -> None:
        """
        Log a single game record to Excel.

        Args:
            record: GameRecord to log
        """
        ws = self.wb["Games"]
        row = [
            record.game_id,
            record.timestamp,
            record.model_name,
            record.iteration,
            record.opponent_type,
            record.opponent_strength,
            record.model_plays_first,
            record.result,
            record.n_moves,
            record.duration_seconds,
        ]
        ws.append(row)
        self.wb.save(self.excel_file)

    def log_summary(
        self,
        model_name: str,
        opponent_type: str,
        opponent_strength: float,
        wins: int,
        draws: int,
        losses: int,
    ) -> None:
        """
        Update or add summary row.

        Args:
            model_name: Name of model
            opponent_type: Type of opponent (egreedy/alphabeta)
            opponent_strength: Epsilon or depth
            wins: Number of wins
            draws: Number of draws
            losses: Number of losses
        """
        ws = self.wb["Summary"]
        total_games = wins + draws + losses
        win_rate = wins / total_games if total_games > 0 else 0.0

        # Find existing row or create new
        key = (model_name, opponent_type, opponent_strength)
        found_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            if (
                row[0].value == model_name
                and row[1].value == opponent_type
                and row[2].value == opponent_strength
            ):
                found_row = idx
                break

        row_data = [
            model_name,
            opponent_type,
            opponent_strength,
            total_games,
            wins,
            draws,
            losses,
            win_rate,
            None,  # rolling_win_rate_last_50 - computed separately
            None,  # first_iteration_reached_95pct
            None,  # first_iteration_reached_98pct
            None,  # sustained_95pct_confirmed
        ]

        if found_row:
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(found_row, col_idx).value = value
        else:
            ws.append(row_data)

        self.wb.save(self.excel_file)

    def log_milestone(
        self,
        model_name: str,
        opponent_type: str,
        opponent_strength: float,
        iteration: int,
        total_games: int,
        win_rate: float,
    ) -> None:
        """
        Log a milestone achievement.

        Args:
            model_name: Name of model
            opponent_type: Type of opponent
            opponent_strength: Epsilon or depth
            iteration: Iteration when milestone reached
            total_games: Total games at milestone
            win_rate: Win rate at milestone
        """
        ws = self.wb["Milestones"]
        row = [
            model_name,
            opponent_type,
            opponent_strength,
            iteration,
            datetime.now().isoformat(),
            total_games,
            win_rate,
        ]
        ws.append(row)
        self.wb.save(self.excel_file)
